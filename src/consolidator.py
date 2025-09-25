#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ESG資料彙整模組 v1.0
將多個Excel檔案彙整成一個總覽表
"""

import os
import re
import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ESGDataConsolidator:
    """
    ESG資料彙整器
    
    功能：
    - 自動掃描results目錄下的所有Excel檔案
    - 排除檔名包含'無提取'的檔案
    - 從檔名中提取標準化的公司名稱
    - 按年度和公司分組生成彙整報告
    - 包含統計摘要和美化格式
    """
    
    def __init__(self, results_path: str):
        self.results_path = Path(results_path)
        
        print(f"📊 初始化ESG資料彙整器")
        print(f"📁 結果目錄: {self.results_path}")
        print(f"⚠️ 注意：檔名包含'無提取'的檔案將被自動排除")
        print(f"🏢 優先使用檔名中的公司資訊（支援標準化格式）")
    
    def consolidate_all_results(self) -> str:
        """彙整所有結果到一個Excel檔案，排除'無提取'檔案"""
        print("\n🚀 開始彙整所有ESG提取結果...")
        print("=" * 60)
        
        # 1. 掃描並分析所有Excel檔案（排除'無提取'）
        excel_files = self._scan_excel_files()
        if not excel_files:
            print("❌ 未找到任何有效的Excel結果檔案")
            print("💡 提示：包含'無提取'的檔案已自動排除")
            return None
        
        print(f"📄 找到 {len(excel_files)} 個有效Excel結果檔案（已排除'無提取'檔案）")
        
        # 2. 解析檔案信息
        parsed_files = self._parse_file_info(excel_files)
        print(f"✅ 成功解析 {len(parsed_files)} 個檔案")
        
        # 3. 載入所有資料
        all_data = self._load_all_data(parsed_files)
        print(f"📚 載入完成，共 {len(all_data)} 筆資料")
        
        # 4. 生成彙整報告
        output_path = self._create_consolidated_excel(all_data, parsed_files)
        
        print(f"✅ 彙整完成！")
        print(f"📊 輸出檔案: {output_path}")
        
        return output_path
    
    def _scan_excel_files(self) -> List[Path]:
        """掃描所有Excel檔案，排除包含'無提取'的檔案"""
        excel_files = []
        excluded_files = []
        
        # 查找所有Excel檔案
        patterns = [
            "提取結果_*.xlsx",
            "*平衡版*.xlsx", 
            "*高精度*.xlsx"
        ]
        
        for pattern in patterns:
            files = list(self.results_path.glob(pattern))
            for file in files:
                # 檢查檔名是否包含"無提取"
                if "無提取" in file.name:
                    excluded_files.append(file)
                    print(f"   ⊗ 排除檔案: {file.name} (包含'無提取')")
                else:
                    excel_files.append(file)
        
        # 顯示排除統計
        if excluded_files:
            print(f"📋 排除了 {len(excluded_files)} 個'無提取'檔案")
        
        # 去重並排序
        unique_files = list(set(excel_files))
        unique_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        return unique_files
    
    def _extract_company_from_filename(self, filename: str) -> Tuple[str, str, str]:
        """
        從檔名中提取公司資訊
        
        支援格式：
        - ESG提取結果_4303_信立_2024.xlsx (標準化格式)
        - ESG提取結果_信立_2024.xlsx (舊格式)
        - ESG提取結果_南亞塑膠工業_2023.xlsx (舊格式)
        
        Returns:
            Tuple[股票代號, 公司名稱, 年度]
        """
        # 移除前綴和副檔名
        clean_name = filename.replace('ESG提取結果_', '').replace('.xlsx', '')
        
        stock_code = ""
        company_name = ""
        year = ""
        
        # 提取年度
        year_match = re.search(r'(202[0-9])', clean_name)
        if year_match:
            year = year_match.group(1)
            clean_name = clean_name.replace(f'_{year}', '').replace(year, '')
        
        # 檢查是否為標準化格式（包含股票代號）
        # 格式: 代號_公司名 或 代號_公司名_其他
        stock_code_pattern = r'^(\d{4}A?|[A-Z]+\d+)_(.+?)(?:_.*)?$'
        stock_match = re.match(stock_code_pattern, clean_name.strip('_'))
        
        if stock_match:
            # 標準化格式
            stock_code = stock_match.group(1)
            company_name = stock_match.group(2)
            print(f"   📊 標準化格式: {stock_code} - {company_name}")
        else:
            # 舊格式，直接使用公司名稱
            company_name = clean_name.strip('_')
            print(f"   📄 舊格式: {company_name}")
        
        # 清理公司名稱
        if company_name:
            # 移除常見後綴
            suffixes_to_remove = [
                "股份有限公司", "有限公司", "公司", 
                "工業", "化學", "塑膠", "科技"
            ]
            for suffix in suffixes_to_remove:
                if company_name.endswith(suffix):
                    company_name = company_name[:-len(suffix)].strip()
                    break
        
        return stock_code, company_name, year
    
    def _parse_file_info(self, excel_files: List[Path]) -> List[Dict]:
        """解析檔案信息，優先使用檔名中的公司資訊"""
        parsed_files = []
        
        for file_path in excel_files:
            try:
                filename = file_path.stem
                
                print(f"📄 解析檔案: {filename}")
                
                # 從檔名提取公司資訊
                stock_code, filename_company, filename_year = self._extract_company_from_filename(filename)
                
                # 解析版本
                if "平衡版" in filename:
                    version = "平衡版"
                elif "高精度" in filename:
                    version = "高精度版"
                else:
                    version = "標準版"
                
                # 嘗試從Excel內容中讀取更詳細的信息作為備用
                excel_company, excel_year = self._extract_company_info_from_excel(file_path)
                
                # 決定最終使用的公司名稱和年度
                final_company = filename_company if filename_company else excel_company
                final_year = filename_year if filename_year else excel_year
                
                # 如果有股票代號，加入到公司名稱中
                if stock_code:
                    display_company = f"{stock_code} {final_company}"
                else:
                    display_company = final_company
                
                parsed_files.append({
                    'file_path': file_path,
                    'filename': filename,
                    'stock_code': stock_code,
                    'company_name': display_company,  # 用於顯示的完整名稱
                    'company_name_only': final_company,  # 僅公司名稱
                    'report_year': final_year,
                    'version': version,
                    'file_time': datetime.fromtimestamp(file_path.stat().st_mtime),
                    'source': 'filename' if filename_company else 'excel_content'
                })
                
                print(f"   ✓ {display_company} - {final_year} ({version}) [來源: {'檔名' if filename_company else 'Excel內容'}]")
                
            except Exception as e:
                print(f"   ⚠️ 解析失敗 {file_path.name}: {e}")
                continue
        
        return parsed_files
    
    def _extract_company_info_from_excel(self, file_path: Path) -> Tuple[str, str]:
        """從Excel檔案中提取公司名稱和年度（作為備用）"""
        try:
            # 讀取第一個工作表的前幾行
            df = pd.read_excel(file_path, nrows=5)
            
            company_name = "未知公司"
            report_year = ""
            
            # 查找公司信息
            for col in df.columns:
                for idx, row in df.iterrows():
                    cell_value = str(row[col])
                    
                    # 提取公司名稱
                    if "公司:" in cell_value:
                        company_match = re.search(r'公司:\s*(.+)', cell_value)
                        if company_match:
                            company_name = company_match.group(1).strip()
                    
                    # 提取報告年度
                    if "報告年度:" in cell_value:
                        year_match = re.search(r'報告年度:\s*(202[0-9])', cell_value)
                        if year_match:
                            report_year = year_match.group(1)
            
            return company_name, report_year
            
        except Exception as e:
            print(f"   ⚠️ 讀取Excel失敗: {e}")
            return "未知公司", ""
    
    def _load_all_data(self, parsed_files: List[Dict]) -> List[Dict]:
        """載入所有資料"""
        all_data = []
        
        for file_info in parsed_files:
            try:
                file_path = file_info['file_path']
                
                # 嘗試讀取不同的工作表名稱
                sheet_names_to_try = [
                    '平衡版提取結果', '提取結果', 'Sheet1', 0
                ]
                
                df = None
                for sheet_name in sheet_names_to_try:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        break
                    except:
                        continue
                
                if df is None:
                    print(f"   ⚠️ 無法讀取 {file_path.name}")
                    continue
                
                # 跳過標題行（前2行通常是公司信息和空行）
                data_start_row = 0
                for idx, row in df.iterrows():
                    if idx < 5:  # 檢查前5行
                        first_cell = str(row.iloc[0]) if len(row) > 0 else ""
                        if "公司:" not in first_cell and first_cell.strip() and first_cell != "nan":
                            data_start_row = idx
                            break
                
                if data_start_row < len(df):
                    df = df.iloc[data_start_row:]
                
                # 為每一行添加檔案信息
                for idx, row in df.iterrows():
                    if len(row) > 0 and str(row.iloc[0]).strip() and str(row.iloc[0]) != "nan":
                        data_row = {
                            'stock_code': file_info['stock_code'],
                            'company_name': file_info['company_name'],  # 完整顯示名稱
                            'company_name_only': file_info['company_name_only'],  # 僅公司名稱
                            'report_year': file_info['report_year'],
                            'version': file_info['version'],
                            'file_name': file_info['filename'],
                            'source_file': file_path.name,
                            'info_source': file_info['source']
                        }
                        
                        # 添加原始數據列
                        for col_idx, col_name in enumerate(df.columns):
                            if col_idx < len(row):
                                data_row[col_name] = row.iloc[col_idx]
                        
                        all_data.append(data_row)
                
                print(f"   ✓ 載入 {file_info['company_name']} - {file_info['report_year']} ({len(df)} 筆)")
                
            except Exception as e:
                print(f"   ❌ 載入失敗 {file_info['file_path'].name}: {e}")
                continue
        
        return all_data
    
    def _create_consolidated_excel(self, all_data: List[Dict], parsed_files: List[Dict]) -> str:
        """創建彙整的Excel檔案"""
        output_filename = f"ESG彙整報告.xlsx"
        output_path = self.results_path / output_filename
        
        print(f"📊 生成彙整Excel: {output_filename}")
        
        # 轉換為DataFrame
        df_all = pd.DataFrame(all_data)
        
        if df_all.empty:
            print("❌ 沒有有效數據可彙整")
            return None
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 按年度分組的工作表
            years = sorted(set(item['report_year'] for item in all_data if item['report_year'] and item['report_year'] != "未知年度"), reverse=True)
            
            for year in years:
                year_data = [item for item in all_data if item['report_year'] == year]
                if year_data:
                    year_df = pd.DataFrame(year_data)
                    sheet_name = f"{year}年總覽"
                    year_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"   ✓ 創建工作表: {sheet_name} ({len(year_data)} 筆)")
            
            # 按公司分組的工作表（使用完整顯示名稱）
            companies = sorted(set(item['company_name'] for item in all_data if item['company_name'] and item['company_name'] != "未知公司"))
            
            for company in companies:
                company_data = [item for item in all_data if item['company_name'] == company]
                if company_data:
                    company_df = pd.DataFrame(company_data)
                    # 按年度排序
                    company_df = company_df.sort_values('report_year', ascending=False)
                    
                    # 清理公司名稱作為工作表名稱（移除股票代號中的特殊字符）
                    safe_company_name = re.sub(r'[^\w\s-]', '', company).strip()[:25]  # Excel工作表名稱限制
                    sheet_name = f"{safe_company_name}總覽"
                    
                    company_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"   ✓ 創建工作表: {sheet_name} ({len(company_data)} 筆)")
            
            # 總覽摘要工作表
            summary_data = self._create_summary_data(parsed_files, all_data)
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='彙整摘要', index=False)
            print(f"   ✓ 創建工作表: 彙整摘要")
        
        # 美化Excel格式
        self._format_excel(output_path)
        
        return str(output_path)
    
    def _create_summary_data(self, parsed_files: List[Dict], all_data: List[Dict]) -> List[Dict]:
        """創建摘要數據"""
        summary_data = []
        
        # 總體統計
        total_files = len(parsed_files)
        total_records = len(all_data)
        companies = set(item['company_name'] for item in all_data if item['company_name'] and item['company_name'] != "未知公司")
        years = set(item['report_year'] for item in all_data if item['report_year'] and item['report_year'] != "未知年度")
        
        # 檔名來源統計
        filename_source_count = len([f for f in parsed_files if f['source'] == 'filename'])
        excel_source_count = len([f for f in parsed_files if f['source'] == 'excel_content'])
        
        # 按公司統計
        company_stats = {}
        for item in all_data:
            company = item['company_name']
            if company and company != "未知公司":
                if company not in company_stats:
                    company_stats[company] = {
                        'years': set(), 
                        'records': 0, 
                        'stock_code': item.get('stock_code', ''),
                        'info_source': item.get('info_source', '')
                    }
                company_stats[company]['years'].add(item['report_year'])
                company_stats[company]['records'] += 1
        
        # 按年度統計
        year_stats = {}
        for item in all_data:
            year = item['report_year']
            if year and year != "未知年度":
                if year not in year_stats:
                    year_stats[year] = {'companies': set(), 'records': 0}
                year_stats[year]['companies'].add(item['company_name'])
                year_stats[year]['records'] += 1
        
        # 生成摘要
        summary_data.append({
            '統計項目': '彙整總覽',
            '數值': f'檔案數: {total_files}, 總筆數: {total_records}',
            '詳細說明': f'涵蓋 {len(companies)} 家公司, {len(years)} 個年度',
            '彙整時間': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # 資訊來源統計
        summary_data.append({
            '統計項目': '資訊來源',
            '數值': f'檔名: {filename_source_count}, Excel內容: {excel_source_count}',
            '詳細說明': f'檔名標準化率: {filename_source_count/total_files*100:.1f}%' if total_files > 0 else '0%',
            '彙整時間': ''
        })
        
        summary_data.append({'統計項目': '', '數值': '', '詳細說明': '', '彙整時間': ''})
        
        # 公司統計（顯示股票代號和資訊來源）
        summary_data.append({
            '統計項目': '公司統計', 
            '數值': '年度範圍', 
            '詳細說明': '提取筆數', 
            '彙整時間': '股票代號/來源'
        })
        
        for company, stats in company_stats.items():
            years_range = f"{min(stats['years'])}-{max(stats['years'])}" if len(stats['years']) > 1 else list(stats['years'])[0]
            info_detail = f"{stats['stock_code']}" if stats['stock_code'] else f"來源:{stats['info_source']}"
            
            summary_data.append({
                '統計項目': company,
                '數值': years_range,
                '詳細說明': f"{stats['records']} 筆",
                '彙整時間': info_detail
            })
        
        summary_data.append({'統計項目': '', '數值': '', '詳細說明': '', '彙整時間': ''})
        
        # 年度統計
        summary_data.append({
            '統計項目': '年度統計', 
            '數值': '公司數量', 
            '詳細說明': '提取筆數', 
            '彙整時間': '公司清單'
        })
        
        for year in sorted(year_stats.keys(), reverse=True):
            stats = year_stats[year]
            # 只顯示公司名稱的核心部分，不包含股票代號
            company_names = [item.get('company_name_only', item['company_name']) for item in all_data if item['report_year'] == year]
            company_list = ', '.join(sorted(set(company_names)))
            
            summary_data.append({
                '統計項目': f"{year}年",
                '數值': f"{len(stats['companies'])} 家公司",
                '詳細說明': f"{stats['records']} 筆",
                '彙整時間': company_list[:100] + ('...' if len(company_list) > 100 else '')
            })
        
        return summary_data
    
    def _format_excel(self, output_path: str):
        """美化Excel格式"""
        try:
            workbook = openpyxl.load_workbook(output_path)
            
            # 定義樣式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # 設置標題行格式
                if worksheet.max_row > 0:
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = border
                    
                    # 自動調整列寬
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(output_path)
            print(f"✨ Excel格式美化完成")
            
        except Exception as e:
            print(f"⚠️ Excel格式化失敗: {e}")

def consolidate_esg_results(results_path: str) -> str:
    """彙整ESG結果的主函數"""
    consolidator = ESGDataConsolidator(results_path)
    return consolidator.consolidate_all_results()

# =============================================================================
# 測試功能
# =============================================================================

def test_filename_parsing():
    """測試檔名解析功能"""
    print("🧪 測試Excel檔名解析")
    print("=" * 50)
    
    # 測試檔名
    test_filenames = [
        "ESG提取結果_4303_信立_2024.xlsx",
        "ESG提取結果_1303_南亞_2023.xlsx",
        "ESG提取結果_4306_炎洲_2024.xlsx",
        "ESG提取結果_信立_2024.xlsx",  # 舊格式
        "ESG提取結果_南亞塑膠工業_2023.xlsx",  # 舊格式
        "ESG提取結果_平衡版_台塑_2024.xlsx",
    ]
    
    consolidator = ESGDataConsolidator("./test")
    
    for filename in test_filenames:
        print(f"\n📄 測試檔名: {filename}")
        stock_code, company_name, year = consolidator._extract_company_from_filename(filename)
        
        if stock_code:
            display_name = f"{stock_code} {company_name}"
        else:
            display_name = company_name
            
        print(f"   結果: {display_name} - {year}")

def test_consolidation():
    """測試彙整功能"""
    print("🧪 測試ESG資料彙整功能")
    
    # 假設有一些測試檔案
    test_results_path = "./test_results"
    
    try:
        consolidator = ESGDataConsolidator(test_results_path)
        result_path = consolidator.consolidate_all_results()
        
        if result_path:
            print(f"✅ 測試成功: {result_path}")
        else:
            print("❌ 測試失敗")
            
    except Exception as e:
        print(f"❌ 測試錯誤: {e}")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        if sys.argv[1] == "--test-parsing":
            test_filename_parsing()
        elif sys.argv[1] == "--test-consolidation":
            test_consolidation()
        else:
            print("用法:")
            print("  python consolidator.py --test-parsing       # 測試檔名解析")
            print("  python consolidator.py --test-consolidation # 測試彙整功能")
    else:
        test_filename_parsing()